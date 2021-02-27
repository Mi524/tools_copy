from django.http import HttpResponse, HttpResponseRedirect, FileResponse
from django.urls import reverse
from django.template import loader
from django.shortcuts import render

from .models import BasicCountry, InputData, UndertakeTeam

from openpyxl import load_workbook, Workbook
from openpyxl.writer.excel import save_virtual_workbook

from sqlalchemy import create_engine

# 转换日期格式
from django.db.models import F, Func, Value, CharField

# 导出写入不成功的记录
import io

# 首页
app_name = "basic_tables"


def write_failure_records(fail_header, fail_records):
    # 导入失败的记录
    wb = Workbook()
    wb.create_sheet()
    ws = wb.active
    # 写入表头
    ws.append(fail_header)
    for f in fail_records:
        ws.append(f)
    wb.close()
    # 导出失败的记录到EXCEL
    response = HttpResponse(content=save_virtual_workbook(
        wb), content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename=Data fail to import.xlsx'
    # 无法返回多个response，保存之后刷新获取导入结果
    return response


def index(request):
    input_data_list = InputData.objects.all().order_by(
        'region', 'start_date', 'test_model')

    return render(request, "basic_tables/index.html",
                  context={'data': input_data_list,
                           'data_count': input_data_list.count()})


def upload_inputdata(request):
    fail_records = []
    if request.method == 'POST':
        print(request)
        # 处理输入的EXCEL，将EXCEL数据导入进数据库
        excel_file = request.FILES['excel_file']
        try:
            wb = load_workbook(excel_file)
        except:
            return HttpResponseRedirect(reverse("basic_tables:index"))
        # 读取默认的第一个sheet
        ws = wb.worksheets[0]

        # 先检查一遍是否有日期问题的数据
        row_counter = 0
        for row in list(ws.iter_rows())[1:]:
            row_counter += 1
            row_data = []
            for cell in row:
                row_data.append(cell.value)

            try:
                p = InputData.objects.create(
                    test_model=row_data[0],
                    type_of_test=row_data[1],
                    details=row_data[2],
                    region=row_data[3],
                    start_date=row_data[4],
                    end_date=row_data[5],
                    person_day=row_data[6],
                    remark=row_data[7],
                    engineer=row_data[8],
                    update_reason=row_data[9],
                    platform_info=row_data[10],
                    status=row_data[11],
                    bpm_link=row_data[12]
                )
            except:
                # 生成失败的记录往底下放
                fail_records.append(row_data)

        # 如果有存在无法写入的记录，写入EXCEL表格导出
        fail_header = [
            "测试机型\nProject",
            "业务模块\nType of test",
            "测试项\nDetails",
            "国家/地区\nCountry/Region",
            "计划开始时间\nStart time",
            "计划完成时间\nEnd time",
            "人*天\npersons*day",
            "备注\nRemarks",
            "跟机工程师\nEngineer",
            "新增/变更原因\nUpdate Reason",
            "平台信息\nPlatform information",
            "完成情况\nCompletion progress",
            "BPM测试单链接\ntest link"
        ]

        if fail_records:
            return write_failure_records(fail_header, fail_records)
    return HttpResponseRedirect(reverse('basic_tables:index'))


def clear_inputdata(request):
    # 清空表内的所有数据
    InputData.objects.all().delete()
    return HttpResponseRedirect(reverse("basic_tables:index"))


def info_team(request):
    # 存储承接团队表的页面
    object_list = UndertakeTeam.objects.all().order_by(
        'country_cn', 'undertake_date', 'undertake_team')
    return render(request, 'basic_tables/info_team.html',
                  context={'data': object_list,
                           'data_count': object_list.count()})


# 承接团队的基础表
def upload_team(request):
    fail_records = []
    if request.method == 'POST':
        # 处理输入的EXCEL，将EXCEL数据导入进数据库
        excel_file = request.FILES['excel_file']
        try:
            wb = load_workbook(excel_file)
        except:
            return HttpResponseRedirect(reverse("basic_tables:info_team"))
        # 读取默认的第一个sheet
        ws = wb.worksheets[0]

        # 先检查一遍是否有日期问题的数据
        row_counter = 0
        for row in list(ws.iter_rows())[1:]:
            row_counter += 1
            row_data = []
            for cell in row:
                row_data.append(cell.value)

            # try:
            p = UndertakeTeam.objects.create(
                country_cn=row_data[0],
                undertake_team=row_data[1],
                undertake_date=row_data[2]
            )
        # except:
            # 生成失败的记录往底下放
            fail_records.append(row_data)

        # 如果有存在无法写入的记录，写入EXCEL表格导出
        fail_header = [
            "国家/地区",
            "承接团队",
            "更新日期",
        ]

        if fail_records:
            return write_failure_records(fail_header, fail_records)

    return HttpResponseRedirect(reverse('basic_tables:info_team'))


def data_basic_country(request):
    input_country_list = BasicCountry.objects.order_by('id')
    output = ','.join([i for i in input_country_list])

    return render(request, 'basic_tables/basic_country.html', context={})
